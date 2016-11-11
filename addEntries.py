
import openpyxl

while True:
    newEntry = input('Add new expense? \Y or N? \n')
    if newEntry == 'y':
        entryDate = input('Enter purchase date \n')
        entryLocation = input('Enter purchase location \n')
        entryAmount = input('Enter cost \n')
        entryCategory = input('Enter purchase category \n')

        wb = openpyxl.load_workbook('Expenses.xlsx')
        type(wb)
        ws = wb.worksheets[0]
        ws['A2'] = entryDate
        ws['B2'] = entryLocation
        ws['C2'] = entryAmount
        ws['D2'] = entryCategory
        wb.save('Expenses.xlsx')

    else:
        break